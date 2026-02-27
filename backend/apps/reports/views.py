import os
import threading
import asyncio
from asgiref.sync import sync_to_async
from django.shortcuts import get_object_or_404
from django.db.models import Q
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from channels.layers import get_channel_layer
import io
from io import BytesIO
import json
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from .models import Report, RequestStats
from .serializers import (
    ReportListSerializer,
    ReportDetailSerializer,
    ReportCreateSerializer,
    ReportRunSerializer,
)
from .locust_engine import locust_engine


class ReportListCreateView(generics.ListCreateAPIView):
    """报告列表/创建"""

    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == "POST":
            return ReportCreateSerializer
        return ReportListSerializer

    def get_queryset(self):
        queryset = Report.objects.filter(created_by=self.request.user)

        # 搜索
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(description__icontains=search)
            )

        # 状态过滤
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # 场景过滤
        scenario_id = self.request.query_params.get("scenario")
        if scenario_id:
            queryset = queryset.filter(scenario_id=scenario_id)

        return queryset.select_related("scenario")

    def list(self, request, *args, **kwargs):
        """重写 list 方法，返回统一格式"""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)

        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return Response(
                {
                    "code": 0,
                    "message": "success",
                    "data": {
                        "results": serializer.data,
                        "count": self.paginator.page.paginator.count,
                        "page": self.paginator.page.number,
                        "page_size": self.paginator.page.paginator.per_page,
                    },
                }
            )

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "code": 0,
                "message": "success",
                "data": {"results": serializer.data, "count": len(serializer.data)},
            }
        )

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        # 返回创建的报告详情
        report = Report.objects.get(id=serializer.instance.id)
        detail_serializer = ReportDetailSerializer(report)

        return Response(
            {"code": 0, "message": "创建成功", "data": detail_serializer.data},
            status=status.HTTP_201_CREATED,
        )


class ReportDetailView(generics.RetrieveDestroyAPIView):
    """报告详情/删除"""

    serializer_class = ReportDetailSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = "pk"

    def get_queryset(self):
        return Report.objects.filter(created_by=self.request.user)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({"code": 0, "message": "success", "data": serializer.data})

    def perform_destroy(self, instance):
        # 删除关联文件
        stats_file = instance.stats_file
        if stats_file and os.path.exists(stats_file):
            try:
                os.remove(stats_file)
            except:
                pass
        instance.delete()


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def run_load_test(request, pk):
    """运行压测"""
    report = get_object_or_404(Report, pk=pk, created_by=request.user)

    if report.status == "running":
        return Response(
            {"code": 400, "message": "压测正在运行中"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 更新状态
    report.status = "running"
    from django.utils import timezone

    report.started_at = timezone.now()
    report.save()

    # 生成 Locust 文件
    locust_file = locust_engine.generate_locustfile(report.scenario, report.id)

    # 发送WebSocket消息的辅助函数
    def send_websocket_message(
        report_id, message, status=None, stats=None, progress=None
    ):
        """发送WebSocket消息的辅助函数"""
        channel_layer = get_channel_layer()

        data = {
            "type": "load_test_update",
            "message": message,
            "timestamp": timezone.now().isoformat(),
        }

        if status:
            data["status"] = status
        if stats:
            data["stats"] = stats
        if progress is not None:
            data["progress"] = progress

        async def send_message():
            await channel_layer.group_send(f"load_test_{report_id}", data)

        try:
            asyncio.run(send_message())
        except Exception as e:
            print(f"Failed to send WebSocket message: {e}")

    # 在新线程中运行压测
    def run_test():
        try:
            # 发送初始状态
            send_websocket_message(
                str(report.id), "压测已启动", status="running", progress=0
            )

            # 运行压测
            locust_engine.run_load_test(report, locust_file)

            # 压测完成后解析结果
            stats = locust_engine.parse_csv_stats(report.id)

            # 更新报告统计
            total_requests = 0
            total_failures = 0

            for name, data in stats.get("requests", {}).items():
                total_requests += data["num_requests"]
                total_failures += data["num_failures"]

                # 创建请求统计
                RequestStats.objects.create(
                    report=report,
                    request_name=name,
                    method="GET",  # 简化处理
                    url=name,
                    num_requests=data["num_requests"],
                    num_failures=data["num_failures"],
                    avg_response_time=data["avg_response_time"],
                    min_response_time=data["min_response_time"],
                    max_response_time=data["max_response_time"],
                    p50_response_time=data["p50"],
                    p90_response_time=data["p90"],
                    p95_response_time=data["p95"],
                    p99_response_time=data["p99"],
                )

            # 更新报告
            report.status = "completed"
            report.total_requests = total_requests
            report.total_failures = total_failures
            report.success_rate = (
                (total_requests - total_failures) / total_requests * 100
                if total_requests > 0
                else 0
            )
            report.ended_at = timezone.now()
            report.save()

            # 发送完成状态
            send_websocket_message(
                str(report.id),
                "压测已完成",
                status="completed",
                stats=stats,
                progress=100,
            )

        except Exception as e:
            report.status = "failed"
            report.ended_at = timezone.now()
            report.save()

            # 发送错误状态
            send_websocket_message(
                str(report.id), f"压测失败: {str(e)}", status="failed", progress=100
            )

            print(f"Load test failed: {e}")

    thread = threading.Thread(target=run_test)
    thread.daemon = True
    thread.start()

    return Response(
        {"code": 0, "message": "压测已启动", "data": {"report_id": str(report.id)}}
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_load_test_stats(request, pk):
    """获取压测实时统计"""
    report = get_object_or_404(Report, pk=pk, created_by=request.user)

    # 解析 CSV 统计
    stats = locust_engine.parse_csv_stats(report.id)

    return Response(
        {
            "code": 0,
            "message": "success",
            "data": {
                "status": report.status,
                "total_requests": report.total_requests,
                "total_failures": report.total_failures,
                "success_rate": report.success_rate,
                "avg_response_time": report.avg_response_time,
                "rps": report.rps,
                "requests": stats.get("requests", {}),
                "history": stats.get("history", []),
            },
        }
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def compare_reports(request):
    """对比报告"""
    report_ids = request.data.get("report_ids", [])

    if len(report_ids) < 2:
        return Response(
            {"code": 400, "message": "请选择至少两个报告进行对比"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    reports = Report.objects.filter(
        id__in=report_ids, created_by=request.user, status="completed"
    )

    if reports.count() < 2:
        return Response(
            {"code": 400, "message": "无效的报告ID"}, status=status.HTTP_400_BAD_REQUEST
        )

    comparison_data = []
    for report in reports:
        comparison_data.append(
            {
                "id": str(report.id),
                "name": report.name,
                "created_at": report.created_at,
                "users": report.users,
                "duration": report.duration,
                "total_requests": report.total_requests,
                "success_rate": report.success_rate,
                "avg_response_time": report.avg_response_time,
                "p95_response_time": report.p95_response_time,
                "p99_response_time": report.p99_response_time,
                "rps": report.rps,
            }
        )

    return Response({"code": 0, "message": "success", "data": comparison_data})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_report_pdf(request, pk):
    """导出报告为PDF"""
    try:
        report = get_object_or_404(Report, pk=pk, created_by=request.user)

        # 创建PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # 标题
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            spaceAfter=30,
        )
        story.append(Paragraph(f"压测报告: {report.name}", title_style))
        story.append(Spacer(1, 12))

        # 基本信息
        basic_info = [
            ["报告名称", report.name],
            ["场景", report.scenario.name if report.scenario else "N/A"],
            ["并发用户数", str(report.users)],
            ["每秒生成数", str(report.spawn_rate)],
            ["压测时长", f"{report.duration} 秒"],
            ["开始时间", str(report.started_at) if report.started_at else "N/A"],
            ["结束时间", str(report.ended_at) if report.ended_at else "N/A"],
            ["状态", report.get_status_display()],
        ]

        basic_table = Table(basic_info)
        basic_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), "#f0f0f0"),
                    ("TEXTPADDING", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 1, "#cccccc"),
                ]
            )
        )
        story.append(basic_table)
        story.append(Spacer(1, 20))

        # 总体统计
        story.append(Paragraph("总体统计", styles["Heading2"]))
        summary_info = [
            ["总请求数", str(report.total_requests or 0)],
            [
                "成功请求数",
                str((report.total_requests or 0) - (report.total_failures or 0)),
            ],
            ["失败请求数", str(report.total_failures or 0)],
            [
                "成功率",
                f"{report.success_rate:.2f}%" if report.success_rate else "0.00%",
            ],
            [
                "平均响应时间",
                f"{report.avg_response_time:.2f}ms"
                if report.avg_response_time
                else "N/A",
            ],
            ["RPS", f"{report.rps:.2f}" if report.rps else "N/A"],
            [
                "P50响应时间",
                f"{report.p50_response_time:.2f}ms"
                if report.p50_response_time
                else "N/A",
            ],
            [
                "P90响应时间",
                f"{report.p90_response_time:.2f}ms"
                if report.p90_response_time
                else "N/A",
            ],
            [
                "P95响应时间",
                f"{report.p95_response_time:.2f}ms"
                if report.p95_response_time
                else "N/A",
            ],
            [
                "P99响应时间",
                f"{report.p99_response_time:.2f}ms"
                if report.p99_response_time
                else "N/A",
            ],
        ]

        summary_table = Table(summary_info)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), "#f0f0f0"),
                    ("TEXTPADDING", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 1, "#cccccc"),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # 请求详细统计
        story.append(Paragraph("请求详细统计", styles["Heading2"]))
        request_stats = RequestStats.objects.filter(report=report)

        if request_stats.exists():
            request_headers = [
                "请求名称",
                "方法",
                "请求数",
                "失败数",
                "平均响应时间(ms)",
                "P95响应时间(ms)",
                "P99响应时间(ms)",
            ]
            request_data = [request_headers]

            for stat in request_stats:
                request_data.append(
                    [
                        stat.request_name,
                        stat.method,
                        str(stat.num_requests or 0),
                        str(stat.num_failures or 0),
                        f"{stat.avg_response_time:.2f}"
                        if stat.avg_response_time
                        else "N/A",
                        f"{stat.p95_response_time:.2f}"
                        if stat.p95_response_time
                        else "N/A",
                        f"{stat.p99_response_time:.2f}"
                        if stat.p99_response_time
                        else "N/A",
                    ]
                )

            request_table = Table(request_data)
            request_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), "#cccccc"),
                        ("TEXTPADDING", (0, 0), (-1, -1), 6),
                        ("GRID", (0, 0), (-1, -1), 1, "#cccccc"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ]
                )
            )
            story.append(request_table)

        # 构建PDF
        doc.build(story)

        # 返回PDF响应
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="report_{report.id[:8]}_{report.name}.pdf"'
        )
        return response

    except Exception as e:
        import traceback

        print(f"PDF export error: {traceback.format_exc()}")
        return Response(
            {"code": 500, "message": f"导出PDF失败: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_report_excel(request, pk):
    """导出报告为Excel"""
    report = get_object_or_404(Report, pk=pk, created_by=request.user)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"压测报告"

    # 添加标题
    ws["A1"] = f"压测报告: {report.name}"
    ws["A1"].font = ws["A1"].font.copy()
    ws["A1"].font.bold = True
    ws["A1"].font.size = 16

    # 添加基本信息
    ws.append([])
    ws.append(["基本信息"])
    ws.append(["字段", "值"])
    basic_info = [
        ["报告名称", report.name],
        ["场景", report.scenario.name if report.scenario else "N/A"],
        ["并发用户数", report.users or 0],
        ["每秒生成数", report.spawn_rate or 0],
        ["压测时长", f"{report.duration} 秒"],
        ["开始时间", str(report.started_at) if report.started_at else "N/A"],
        ["结束时间", str(report.ended_at) if report.ended_at else "N/A"],
        ["状态", report.get_status_display()],
    ]

    for row in basic_info:
        ws.append(row)

    # 添加总体统计
    ws.append([])
    ws.append([])
    ws.append(["总体统计"])
    ws.append(["字段", "值"])
    summary_info = [
        ["总请求数", report.total_requests or 0],
        ["成功请求数", (report.total_requests or 0) - (report.total_failures or 0)],
        ["失败请求数", report.total_failures or 0],
        ["成功率", f"{report.success_rate:.2f}%" if report.success_rate else "0.00%"],
        [
            "平均响应时间",
            f"{report.avg_response_time:.2f}ms" if report.avg_response_time else "N/A",
        ],
        ["RPS", f"{report.rps:.2f}" if report.rps else "N/A"],
        [
            "P50响应时间",
            f"{report.p50_response_time:.2f}ms" if report.p50_response_time else "N/A",
        ],
        [
            "P90响应时间",
            f"{report.p90_response_time:.2f}ms" if report.p90_response_time else "N/A",
        ],
        [
            "P95响应时间",
            f"{report.p95_response_time:.2f}ms" if report.p95_response_time else "N/A",
        ],
        [
            "P99响应时间",
            f"{report.p99_response_time:.2f}ms" if report.p99_response_time else "N/A",
        ],
    ]

    for row in summary_info:
        ws.append(row)

    # 添加请求详细统计
    request_stats = RequestStats.objects.filter(report=report)
    if request_stats.exists():
        ws.append([])
        ws.append([])
        ws.append(["请求详细统计"])
        ws.append(
            [
                "请求名称",
                "方法",
                "请求数",
                "失败数",
                "平均响应时间(ms)",
                "P95响应时间(ms)",
                "P99响应时间(ms)",
            ]
        )

        for stat in request_stats:
            ws.append(
                [
                    stat.request_name,
                    stat.method,
                    stat.num_requests or 0,
                    stat.num_failures or 0,
                    round(stat.avg_response_time, 2)
                    if stat.avg_response_time
                    else "N/A",
                    round(stat.p95_response_time, 2)
                    if stat.p95_response_time
                    else "N/A",
                    round(stat.p99_response_time, 2)
                    if stat.p99_response_time
                    else "N/A",
                ]
            )

    # 设置列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 返回Excel响应
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="report_{report.id[:8]}_{report.name}.xlsx"'
    )
    return response
